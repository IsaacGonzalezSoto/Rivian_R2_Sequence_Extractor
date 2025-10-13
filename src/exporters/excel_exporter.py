"""
Excel exporter for sequences and transitions data.
Uses openpyxl to create .xlsx files with multiple sheets.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from typing import Dict, Any, List
from ..core.constants import ExcelColors, ExcelFontSizes, MAX_COLUMN_WIDTH, DEFAULT_COLUMN_PADDING


class ExcelExporter:
    """
    Exports data to Excel format (.xlsx) with multiple sheets.
    One file per routine with separate sheets for sequences and transitions.
    """
    
    def __init__(self):
        """Initialize the Excel exporter."""
        # Header styles
        self.header_fill = PatternFill(start_color=ExcelColors.HEADER_FILL, end_color=ExcelColors.HEADER_FILL, fill_type="solid")
        self.header_font = Font(bold=True, color=ExcelColors.HEADER_FONT)
        self.header_alignment = Alignment(horizontal="center", vertical="center")

        # Styles for Complete_Flow sheet
        self.transition_fill = PatternFill(start_color=ExcelColors.TRANSITION_FILL, end_color=ExcelColors.TRANSITION_FILL, fill_type="solid")
        self.transition_font = Font(bold=True, color=ExcelColors.TRANSITION_FONT, size=ExcelFontSizes.TRANSITION)
        self.sequence_fill = PatternFill(start_color=ExcelColors.SEQUENCE_FILL, end_color=ExcelColors.SEQUENCE_FILL, fill_type="solid")
        self.sequence_font = Font(bold=True, color=ExcelColors.SEQUENCE_FONT, size=ExcelFontSizes.SEQUENCE)
        self.step_fill = PatternFill(start_color=ExcelColors.STEP_FILL, end_color=ExcelColors.STEP_FILL, fill_type="solid")
        self.step_font = Font(bold=True, color=ExcelColors.STEP_FONT, size=ExcelFontSizes.STEP)
        self.action_fill = PatternFill(start_color=ExcelColors.ACTION_FILL, end_color=ExcelColors.ACTION_FILL, fill_type="solid")
        self.action_font = Font(bold=True, size=ExcelFontSizes.ACTION)

        # Data cell background - soft beige for eye comfort
        self.data_fill = PatternFill(start_color=ExcelColors.DATA_FILL, end_color=ExcelColors.DATA_FILL, fill_type="solid")
    
    def export(self, sequences_data: Dict[str, Any], transitions_data: Dict[str, Any], digital_inputs_data: Dict[str, Any], actuator_groups_data: Dict[str, Any], output_path: str):
        """
        Export sequences, transitions, digital inputs, and actuator groups to a single Excel file with multiple sheets.

        Args:
            sequences_data: Dictionary with sequence and actuator data
            transitions_data: Dictionary with transition permission data
            digital_inputs_data: Dictionary with digital input tags data
            actuator_groups_data: Dictionary with actuator group tags data
            output_path: Path for the output Excel file
        """
        wb = Workbook()

        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        # Create Complete_Flow sheet (new main view)
        self._create_complete_flow_sheet(wb, sequences_data, transitions_data)

        # Create sequences sheet with actuator group descriptions
        self._create_sequences_sheet(wb, sequences_data, actuator_groups_data)

        # Create transitions sheet if there's data
        if transitions_data and transitions_data.get('transitions'):
            self._create_transitions_sheet(wb, transitions_data)

        # Create digital inputs sheet if there's data
        if digital_inputs_data and digital_inputs_data.get('digital_inputs'):
            self._create_digital_inputs_sheet(wb, digital_inputs_data)

        # Note: Actuator groups are now included in Sequences_Actuators sheet as MM_Group_Description column

        # Save workbook
        wb.save(output_path)
    
    def _create_sequences_sheet(self, wb: Workbook, data: Dict[str, Any], actuator_groups_data: Dict[str, Any] = None):
        """
        Create the Sequences_Actuators sheet.

        Args:
            wb: Workbook object
            data: Sequences data
            actuator_groups_data: Actuator groups data (optional)
        """
        ws = wb.create_sheet("Sequences_Actuators")

        # Create MM mapping dictionary: {MM1: 'Group1 Clamps', MM2: 'Group2 Clamps', ...}
        mm_to_description = {}
        if actuator_groups_data and actuator_groups_data.get('actuator_groups'):
            for group in actuator_groups_data['actuator_groups']:
                mm_to_description[group['tag_name']] = group['description']

        # Headers
        headers = [
            'Routine',
            'Sequence',
            'Step',
            'Action_Index',
            'Action_Name',
            'MM_Number',
            'MM_Group_Description',
            'State',
            'Actuator_Count',
            'Actuators',
            'Validation_Status',
            'Missing_Indices'
        ]
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.header_alignment
        
        # Write data
        row_num = 2
        routine_name = data['routine_name']
        
        for sequence in data['sequences']:
            seq_idx = sequence['sequence_index']
            
            for step in sequence['steps']:
                step_idx = step['step_index']
                
                for action in step['actions']:
                    # Determine validation status
                    validation_status = 'N/A'
                    missing_indices = ''
                    
                    if 'validation' in action:
                        val = action['validation']
                        if val['is_valid'] == True:
                            validation_status = 'OK'
                        elif val['is_valid'] == False:
                            validation_status = 'WARNING'
                            missing_indices = ','.join(map(str, val['missing_indices']))
                        elif val['is_valid'] is None:
                            validation_status = 'UNKNOWN'
                    
                    # Format state
                    state_formatted = ''
                    if action['state']:
                        state_formatted = f"TO {action['state'].upper()}"
                    
                    # Get MM group description
                    mm_number = action['mm_number'] or ''
                    mm_description = mm_to_description.get(mm_number, '') if mm_number else ''

                    # Write one row per actuator
                    if not action['actuators']:
                        # No actuators, write one row
                        row_data = [
                            routine_name,
                            seq_idx,
                            step_idx,
                            action['action_index'],
                            action['action_name'],
                            mm_number,
                            mm_description,
                            state_formatted,
                            action['actuator_count'],
                            '',
                            validation_status,
                            missing_indices
                        ]
                        
                        for col_num, value in enumerate(row_data, 1):
                            cell = ws.cell(row=row_num, column=col_num, value=value)
                            cell.fill = self.data_fill  # Apply beige background
                        row_num += 1
                    else:
                        # One row per actuator
                        for actuator in action['actuators']:
                            row_data = [
                                routine_name,
                                seq_idx,
                                step_idx,
                                action['action_index'],
                                action['action_name'],
                                mm_number,
                                mm_description,
                                state_formatted,
                                action['actuator_count'],
                                actuator['description'],
                                validation_status,
                                missing_indices
                            ]
                            
                            for col_num, value in enumerate(row_data, 1):
                                cell = ws.cell(row=row_num, column=col_num, value=value)
                                cell.fill = self.data_fill  # Apply beige background
                            row_num += 1
        
        # Auto-adjust column widths
        self._adjust_column_widths(ws)
    
    def _create_digital_inputs_sheet(self, wb: Workbook, data: Dict[str, Any]):
        """
        Create the Digital Inputs sheet.
        
        Args:
            wb: Workbook object
            data: Digital inputs data
        """
        ws = wb.create_sheet("Digital Inputs")
        
        # Headers
        headers = [
            'Program',
            'Tag Name',
            'Description',
            'Parent Name',
            'Part Assignment'
        ]
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.header_alignment
        
        # Write data
        row_num = 2
        
        for digital_input in data['digital_inputs']:
            row_data = [
                digital_input['program'],
                digital_input['tag_name'],
                digital_input['description'],
                digital_input['parent_name'],
                digital_input.get('part_assignment', 'N/A')
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.fill = self.data_fill  # Apply beige background
            row_num += 1
        
        # Auto-adjust column widths
        self._adjust_column_widths(ws)

    def _create_transitions_sheet(self, wb: Workbook, data: Dict[str, Any]):
        """
        Create the Transitions sheet.
        
        Args:
            wb: Workbook object
            data: Transitions data
        """
        ws = wb.create_sheet("Transitions")
        
        # Headers
        headers = [
            'Routine',
            'Transition_Index',
            'Permission_Count',
            'Permission_Index',
            'Permission_Value',
            'Comment'
        ]
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.header_alignment
        
        # Write data
        row_num = 2
        routine_name = data['routine_name']
        
        for transition in data['transitions']:
            trans_idx = transition['transition_index']
            permission_count = transition['permission_count']
            
            for permission in transition['permissions']:
                row_data = [
                    routine_name,
                    trans_idx,
                    permission_count,
                    permission['permission_index'],
                    permission['permission_value'],
                    permission['comment']
                ]
                
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.fill = self.data_fill  # Apply beige background
                row_num += 1
        
        # Auto-adjust column widths
        self._adjust_column_widths(ws)
    
    def _adjust_column_widths(self, ws):
        """
        Auto-adjust column widths based on content.
        
        Args:
            ws: Worksheet object
        """
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + DEFAULT_COLUMN_PADDING, MAX_COLUMN_WIDTH)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_complete_flow_sheet(self, wb: Workbook, sequences_data: Dict[str, Any], transitions_data: Dict[str, Any]):
        """
        Create the Complete_Flow sheet showing the hierarchical flow.
        
        Args:
            wb: Workbook object
            sequences_data: Sequences data
            transitions_data: Transitions data
        """
        ws = wb.create_sheet("Complete_Flow", 0)  # Insert as first sheet
        
        # Headers
        headers = ['Type', 'Index', 'Description', 'Details', 'State/Comment']
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.header_alignment
        
        row_num = 2
        routine_name = sequences_data['routine_name']
        
        # Build a mapping of sequence index to transition (if they match)
        transition_map = {}
        if transitions_data and transitions_data.get('transitions'):
            for transition in transitions_data['transitions']:
                trans_idx = transition['transition_index']
                transition_map[trans_idx] = transition
        
        # Process each sequence and its corresponding transition
        for sequence in sequences_data['sequences']:
            seq_idx = sequence['sequence_index']
            
            # Write Transition header if exists
            if seq_idx in transition_map:
                transition = transition_map[seq_idx]
                row_num = self._write_transition_section(ws, row_num, transition)
            
            # Write Sequence header
            row_num = self._write_sequence_section(ws, row_num, sequence)
        
        # Auto-adjust column widths
        self._adjust_column_widths(ws)
    
    def _write_transition_section(self, ws, row_num: int, transition: Dict[str, Any]) -> int:
        """
        Write a transition section with its permissions.
        
        Args:
            ws: Worksheet
            row_num: Current row number
            transition: Transition data
            
        Returns:
            Updated row number
        """
        trans_idx = transition['transition_index']
        perm_count = transition['permission_count']
        
        # Use descriptive name if available, otherwise use default
        if transition.get('transition_name'):
            description = f"Transition State {trans_idx} - {transition['transition_name']}"
        else:
            description = f"Transition {trans_idx}"
        
        # Transition header row
        ws.cell(row=row_num, column=1, value='TRANSITION')
        ws.cell(row=row_num, column=2, value=trans_idx)
        ws.cell(row=row_num, column=3, value=description)
        ws.cell(row=row_num, column=4, value=f'{perm_count} permissions')
        
        # Apply transition header style
        for col in range(1, 6):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = self.transition_fill
            cell.font = self.transition_font
        
        row_num += 1
        
        # Write permissions with beige background
        for permission in transition['permissions']:
            ws.cell(row=row_num, column=1, value='  Permission')
            ws.cell(row=row_num, column=2, value=permission['permission_index'])
            ws.cell(row=row_num, column=3, value=permission['permission_value'])
            ws.cell(row=row_num, column=4, value='')
            ws.cell(row=row_num, column=5, value=permission['comment'])
            
            # Apply beige background to permission rows
            for col in range(1, 6):
                cell = ws.cell(row=row_num, column=col)
                cell.fill = self.data_fill
            
            row_num += 1
        
        # Add blank row for separation
        row_num += 1
        
        return row_num
    
    def _write_sequence_section(self, ws, row_num: int, sequence: Dict[str, Any]) -> int:
        """
        Write a sequence section with its steps, actions, and actuators.
        
        Args:
            ws: Worksheet
            row_num: Current row number
            sequence: Sequence data
            
        Returns:
            Updated row number
        """
        seq_idx = sequence['sequence_index']
        step_count = len(sequence['steps'])
        
        # Use descriptive name if available, otherwise use default
        if sequence.get('sequence_name'):
            description = f"Sequence State {seq_idx} - {sequence['sequence_name']}"
        else:
            description = f"Sequence {seq_idx}"
        
        # Sequence header row
        ws.cell(row=row_num, column=1, value='SEQUENCE')
        ws.cell(row=row_num, column=2, value=seq_idx)
        ws.cell(row=row_num, column=3, value=description)
        ws.cell(row=row_num, column=4, value=f'{step_count} steps')
        
        # Apply sequence header style
        for col in range(1, 6):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = self.sequence_fill
            cell.font = self.sequence_font
        
        row_num += 1
        
        # Write steps
        for step in sequence['steps']:
            row_num = self._write_step_section(ws, row_num, seq_idx, step)
        
        # Add blank row for separation
        row_num += 1
        
        return row_num
    
    def _write_step_section(self, ws, row_num: int, seq_idx: int, step: Dict[str, Any]) -> int:
        """
        Write a step section with its actions and actuators.
        
        Args:
            ws: Worksheet
            row_num: Current row number
            seq_idx: Sequence index
            step: Step data
            
        Returns:
            Updated row number
        """
        step_idx = step['step_index']
        action_count = len(step['actions'])
        
        # Step header row
        ws.cell(row=row_num, column=1, value='  STEP')
        ws.cell(row=row_num, column=2, value=step_idx)
        ws.cell(row=row_num, column=3, value=f'Step {step_idx}')
        ws.cell(row=row_num, column=4, value=f'{action_count} actions')
        
        # Apply step header style
        for col in range(1, 6):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = self.step_fill
            cell.font = self.step_font
        
        row_num += 1
        
        # Write actions
        for action in step['actions']:
            row_num = self._write_action_section(ws, row_num, action)
        
        return row_num
    
    def _write_action_section(self, ws, row_num: int, action: Dict[str, Any]) -> int:
        """
        Write an action section with its actuators.
        
        Args:
            ws: Worksheet
            row_num: Current row number
            action: Action data
            
        Returns:
            Updated row number
        """
        action_idx = action['action_index']
        action_name = action['action_name']
        mm_number = action['mm_number'] or 'N/A'
        
        # Format state
        state_formatted = ''
        if action['state']:
            state_formatted = f"TO {action['state'].upper()}"
        
        # Build details with validation information
        actuator_count = action['actuator_count']
        details = f'{mm_number} - {actuator_count} actuators'
        
        # Add validation status if available
        if 'validation' in action and action['validation']:
            val = action['validation']
            if val['is_valid'] == True:
                array_dim = val.get('array_dimension', actuator_count)
                details += f' [OK: {actuator_count}/{array_dim}]'
            elif val['is_valid'] == False:
                array_dim = val.get('array_dimension', 0)
                missing = ','.join(map(str, val.get('missing_indices', [])))
                details += f' [WARNING: {actuator_count}/{array_dim} - Missing: {missing}]'
            elif val['is_valid'] is None:
                details += f' [UNKNOWN]'
        
        # Action header row
        ws.cell(row=row_num, column=1, value='    ACTION')
        ws.cell(row=row_num, column=2, value=action_idx)
        ws.cell(row=row_num, column=3, value=action_name)
        ws.cell(row=row_num, column=4, value=details)
        ws.cell(row=row_num, column=5, value=state_formatted)
        
        # Apply action header style
        for col in range(1, 6):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = self.action_fill
            cell.font = self.action_font
        
        row_num += 1
        
        # Write actuators with beige background
        for actuator in action['actuators']:
            ws.cell(row=row_num, column=1, value='      Actuator')
            ws.cell(row=row_num, column=2, value=actuator['index'])
            ws.cell(row=row_num, column=3, value=actuator['description'])
            ws.cell(row=row_num, column=4, value=mm_number)
            
            # Apply beige background to actuator rows
            for col in range(1, 6):
                cell = ws.cell(row=row_num, column=col)
                cell.fill = self.data_fill
            
            row_num += 1
        
        return row_num