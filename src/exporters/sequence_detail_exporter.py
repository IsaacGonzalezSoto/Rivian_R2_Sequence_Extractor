"""
Sequence Detail exporter for generating the simplified Start Conditions format.
This exporter creates a single-sheet Excel file showing the initial state of all cylinders and sensors.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from typing import Dict, Any, List
from ..core.constants import ExcelColors, ExcelFontSizes, MAX_COLUMN_WIDTH, DEFAULT_COLUMN_PADDING
from ..core.logger import get_logger

logger = get_logger(__name__)


class SequenceDetailExporter:
    """
    Exports sequence detail data showing Start Conditions with cylinders and sensors.
    Creates a single Excel file with one sheet showing the initial state.
    """

    def __init__(self):
        """Initialize the Sequence Detail exporter."""
        # Header styles
        self.header_fill = PatternFill(start_color=ExcelColors.HEADER_FILL, end_color=ExcelColors.HEADER_FILL, fill_type="solid")
        self.header_font = Font(bold=True, color=ExcelColors.HEADER_FONT)
        self.header_alignment = Alignment(horizontal="center", vertical="center")

        # Fixed State style (blue background, WHITE text)
        self.fixed_state_fill = PatternFill(start_color=ExcelColors.TRANSITION_FILL, end_color=ExcelColors.TRANSITION_FILL, fill_type="solid")
        self.fixed_state_font = Font(bold=True, color="FFFFFF", size=ExcelFontSizes.TRANSITION)

        # Transition State style (green background, WHITE text)
        self.transition_state_fill = PatternFill(start_color=ExcelColors.SEQUENCE_FILL, end_color=ExcelColors.SEQUENCE_FILL, fill_type="solid")
        self.transition_state_font = Font(bold=True, color="FFFFFF", size=ExcelFontSizes.SEQUENCE)

        # Data cell background - soft beige for eye comfort
        self.data_fill = PatternFill(start_color=ExcelColors.DATA_FILL, end_color=ExcelColors.DATA_FILL, fill_type="solid")

        # Data cell font - WHITE text for dark background
        self.data_font = Font(color="FFFFFF")

    def export(self, common_data: Dict[str, Any], model_data: Dict[str, Any],
               digital_inputs_data: Dict[str, Any], actuator_groups_data: Dict[str, Any],
               valve_mappings_data: Dict[str, Any], all_actuators_data: Dict[str, Any],
               transitions_data: Dict[str, Any], output_path: str):
        """
        Export sequence detail with validation and error handling.

        The logic is:
        1. Common sequences define the base state (usually all cylinders to HOME)
        2. Model's first sequence may override some cylinder positions
        3. Result = Common state + Model first sequence overrides
        4. All sensors are listed with their part assignments
        5. Transitions are appended with Fixed State, Wait Conditions, and Actions

        Args:
            common_data: Sequences data from Common routine
            model_data: Sequences data from model routine (e.g., R2S)
            digital_inputs_data: Digital inputs data with part assignments
            actuator_groups_data: Actuator groups data for MM descriptions
            valve_mappings_data: Valve mappings data for valve nomenclature
            all_actuators_data: Complete list of ALL actuators from MM routines
            transitions_data: Transitions data with permissions
            output_path: Path for the output Excel file
        """
        # Validate input data
        validation_warnings = []

        if not common_data or not common_data.get('sequences'):
            validation_warnings.append("Missing or empty Common sequences data")

        if not model_data or not model_data.get('sequences'):
            validation_warnings.append("Missing or empty Model sequences data")

        if not all_actuators_data or not all_actuators_data.get('actuators_by_mm'):
            validation_warnings.append("Missing or empty actuators data")

        if not transitions_data or not transitions_data.get('transitions'):
            validation_warnings.append("Missing or empty transitions data")

        if validation_warnings:
            logger.warning("⚠️ DATA VALIDATION WARNINGS:")
            for warning in validation_warnings:
                logger.warning(f"  - {warning}")

        # Check for sequence/transition count consistency
        if model_data and transitions_data:
            num_sequences = len(model_data.get('sequences', []))
            num_transitions = len(transitions_data.get('transitions', []))
            if num_sequences != num_transitions:
                logger.warning(f"⚠️ WARNING: Sequence/Transition count mismatch:")
                logger.warning(f"  - Sequences: {num_sequences}")
                logger.warning(f"  - Transitions: {num_transitions}")

        # Try to export with error handling
        try:
            wb = Workbook()

            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])

            # Create the main Sequence Detail sheet
            self._create_sequence_detail_sheet(
                wb, common_data, model_data, digital_inputs_data,
                actuator_groups_data, valve_mappings_data,
                all_actuators_data, transitions_data
            )

            # Save with error handling
            try:
                wb.save(output_path)
                logger.info(f"✅ Successfully created Sequence Detail: {output_path}")
            except PermissionError:
                raise Exception(
                    f"Cannot write to {output_path} - file may be open in Excel. "
                    "Please close the file and try again."
                )

        except Exception as e:
            logger.error(f"Error creating Sequence Detail export: {str(e)}")
            raise

    def _create_sequence_detail_sheet(self, wb: Workbook, common_data: Dict[str, Any],
                                     model_data: Dict[str, Any], digital_inputs_data: Dict[str, Any],
                                     actuator_groups_data: Dict[str, Any], valve_mappings_data: Dict[str, Any],
                                     all_actuators_data: Dict[str, Any], transitions_data: Dict[str, Any]):
        """
        Create the Sequence Detail sheet with Start Conditions and Transitions.

        Args:
            wb: Workbook object
            common_data: Common sequences data
            model_data: Model sequences data
            digital_inputs_data: Digital inputs data
            actuator_groups_data: Actuator groups data
            valve_mappings_data: Valve mappings data
            all_actuators_data: Complete list of ALL actuators from MM routines
            transitions_data: Transitions data with permissions
        """
        ws = wb.create_sheet("Sequence Detail")

        # Headers with 17 columns (12 original + 5 timing columns)
        headers = [
            'ID',
            'Style',
            'Row Type',
            'Actor Type',
            'Actor/Unit',
            'Custom Description',
            'Custom \nDuration',
            'Standard \nAction/\nStatus',
            'Standard Duration',
            'Actor/Group \nDescription',
            'Actor/\nGroup \nName',
            'Valve \nName',
            'Ind. \nStart',      # Timing column (empty - filled manually)
            'Dep. ID',           # Timing column (empty - filled manually)
            'Start',             # Timing column (empty - filled manually)
            'Duration',          # Timing column (empty - filled manually)
            'End'                # Timing column (empty - filled manually)
        ]

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.header_alignment

        # Build MM description mapping
        mm_to_description = {}
        if actuator_groups_data and actuator_groups_data.get('actuator_groups'):
            for group in actuator_groups_data['actuator_groups']:
                mm_to_description[group['tag_name']] = group['description']

        # Build valve mapping dictionary: {MM1: {manifold: '...', valve_work: '1A', valve_home: '1B'}, ...}
        mm_to_valve = {}
        if valve_mappings_data and valve_mappings_data.get('valve_mappings'):
            mm_to_valve = valve_mappings_data['valve_mappings']

        # Build initial cylinder state from ALL actuators extracted from MM routines
        # Key: (mm_number, index) for uniqueness, Value: {actuator_name, state, mm_number, description, manifold, valve_work, valve_home}
        cylinder_state = {}

        # Initialize ALL cylinders with HOME state
        if all_actuators_data and all_actuators_data.get('actuators_by_mm'):
            for mm_number, actuators in all_actuators_data['actuators_by_mm'].items():
                mm_description = mm_to_description.get(mm_number, '')

                # Get valve mapping information
                manifold = ''
                valve_work = ''
                valve_home = ''
                if mm_number in mm_to_valve:
                    valve_info = mm_to_valve[mm_number]
                    manifold = valve_info.get('manifold', '')
                    valve_work = valve_info.get('valve_work', '')
                    valve_home = valve_info.get('valve_home', '')

                # Add all actuators with default HOME state
                for actuator in actuators:
                    actuator_name = actuator['description']
                    actuator_index = actuator['index']
                    # Use (mm_number, index) as key to ensure uniqueness
                    unique_key = (mm_number, actuator_index)
                    cylinder_state[unique_key] = {
                        'actuator_name': actuator_name,
                        'state': 'HOME',  # Store as simple state, will format as "AT HOME" for start conditions
                        'mm_number': mm_number,
                        'description': mm_description,
                        'manifold': manifold,
                        'valve_work': valve_work,
                        'valve_home': valve_home
                    }

        # Apply overrides from Common sequences
        if common_data and common_data.get('sequences'):
            for sequence in common_data['sequences']:
                for step in sequence['steps']:
                    for action in step['actions']:
                        state = action.get('state', '')
                        state_formatted = self._format_state_robust(state)
                        mm_number = action.get('mm_number', '')

                        for actuator in action.get('actuators', []):
                            actuator_index = actuator['index']
                            unique_key = (mm_number, actuator_index)
                            # Override state if this actuator exists
                            if unique_key in cylinder_state:
                                cylinder_state[unique_key]['state'] = state_formatted

        # Apply overrides from Model's FIRST sequence
        if model_data and model_data.get('sequences'):
            first_sequence = model_data['sequences'][0] if model_data['sequences'] else None
            if first_sequence:
                for step in first_sequence['steps']:
                    for action in step['actions']:
                        state = action.get('state', '')
                        state_formatted = self._format_state_robust(state)
                        mm_number = action.get('mm_number', '')

                        for actuator in action.get('actuators', []):
                            actuator_index = actuator['index']
                            unique_key = (mm_number, actuator_index)
                            # Override state if this actuator exists
                            if unique_key in cylinder_state:
                                cylinder_state[unique_key]['state'] = state_formatted

        # Sort cylinders by MM group and index
        sorted_cylinders = sorted(
            cylinder_state.items(),
            key=lambda x: (
                self._extract_mm_number_from_key(x[0][0]),  # Sort by MM number from key tuple
                x[0][1]  # Then by index
            )
        )

        # Start writing data rows
        row_num = 2
        id_counter = 1

        # Write Start Conditions Header row
        row_data = [
            id_counter,                      # ID
            'Common',                        # Style
            'Start Conditions Header',       # Row Type
            None,                            # Actor Type (empty)
            None,                            # Actor/Unit (empty - filled manually)
            'HomePos',                       # Custom Description
            None, None, None, None, None, None,  # Other columns empty
            None, None, None, None, None     # Timing columns (empty)
        ]
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.fill = self.data_fill
            cell.font = self.data_font
        row_num += 1
        id_counter += 1

        # Write cylinder rows (Start Conditions, CylinderUnits)
        for unique_key, info in sorted_cylinders:
            actuator_name = info['actuator_name']

            # Calculate valve name based on state using helper method
            mm_number = info['mm_number']
            state = info['state']
            valve_name = self._calculate_valve_name(state, mm_number, mm_to_valve)

            # Format state for Start Conditions: use "AT" prefix (not "TO")
            state_formatted = self._format_start_condition_state(state)

            # Format actuator name with = and -
            actor_group_name = self._format_actor_group_name(actuator_name)

            row_data = [
                id_counter,                      # ID
                'Common',                        # Style (always "Common" for start conditions)
                'Start Conditions',              # Row Type
                'CylinderUnits',                 # Actor Type
                None,                            # Actor/Unit (empty - filled manually)
                None,                            # Custom Description (empty)
                None,                            # Custom Duration (empty)
                state_formatted,                 # Standard Action/Status (AT WORK/HOME)
                0.0,                             # Standard Duration
                info['description'],             # Actor/Group Description
                actor_group_name,                # Actor/Group Name (=MM1-MMB1)
                valve_name,                      # Valve Name
                None, None, None, None, None     # Timing columns (empty)
            ]

            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.fill = self.data_fill
                cell.font = self.data_font
            row_num += 1
            id_counter += 1

        # Extract and write sensor rows
        # Only include sensors that have part assignments (not 'N/A')
        if digital_inputs_data and digital_inputs_data.get('digital_inputs'):
            # Group sensors by part assignment
            sensors_by_part = {}
            for sensor in digital_inputs_data['digital_inputs']:
                tag_name = sensor.get('tag_name', '')
                part_assignment = sensor.get('part_assignment', 'N/A')

                # Only include sensors with actual part assignments and starting with BG (part sensors)
                if part_assignment != 'N/A' and tag_name.startswith('BG'):
                    if part_assignment not in sensors_by_part:
                        sensors_by_part[part_assignment] = []
                    sensors_by_part[part_assignment].append(tag_name)

            # Sort by part number using robust part name sorting
            sorted_parts = sorted(sensors_by_part.keys(), key=self._sort_part_name)

            # Write sensor rows
            for part_name in sorted_parts:
                sensors = sorted(sensors_by_part[part_name])
                for sensor_name in sensors:
                    # Format sensor name with = and -
                    sensor_name_formatted = self._format_actor_group_name(sensor_name)

                    row_data = [
                        id_counter,                      # ID
                        'Common',                        # Style
                        'Start Conditions',              # Row Type
                        'SensorUnits',                   # Actor Type
                        None,                            # Actor/Unit (empty - filled manually)
                        None,                            # Custom Description (empty)
                        None,                            # Custom Duration (empty)
                        'OFF',                           # Standard Action/Status (sensors start OFF)
                        0.0,                             # Standard Duration
                        part_name,                       # Actor/Group Description (Part1, Part2, etc.)
                        sensor_name_formatted,           # Actor/Group Name (=BG1-BGB1)
                        None,                            # Valve Name (empty)
                        None, None, None, None, None     # Timing columns (empty)
                    ]

                    for col_num, value in enumerate(row_data, 1):
                        cell = ws.cell(row=row_num, column=col_num, value=value)
                        cell.fill = self.data_fill
                        cell.font = self.data_font
                    row_num += 1
                    id_counter += 1

        # Write transitions section (Fixed State + Wait Conditions + Actions)
        self._write_transitions_section(ws, row_num, id_counter, transitions_data, digital_inputs_data, model_data, actuator_groups_data, valve_mappings_data)

        # Auto-adjust column widths
        self._adjust_column_widths(ws)

    def _extract_mm_number(self, actuator_name: str) -> int:
        """
        Extract MM number from actuator name for sorting.

        Args:
            actuator_name: Actuator name like 'MM1_MMB1', 'MM12_MMB3'

        Returns:
            MM number as integer (e.g., 1, 12)
        """
        import re
        match = re.match(r'MM(\d+)_', actuator_name)
        return int(match.group(1)) if match else 999

    def _extract_mm_number_from_key(self, mm_number: str) -> int:
        """
        Extract MM number from MM tag name for sorting.

        Args:
            mm_number: MM tag name like 'MM1', 'MM12'

        Returns:
            MM number as integer (e.g., 1, 12)
        """
        import re
        match = re.match(r'MM(\d+)', mm_number)
        return int(match.group(1)) if match else 999

    def _extract_kj_name(self, manifold: str) -> str:
        """
        Extract KJ{x} pattern from manifold name.

        Args:
            manifold: Manifold name (e.g., '_010_UA1_KJ1_Hw', '_010UA1KJ1_KEB1_Hw')

        Returns:
            KJ name (e.g., 'KJ1') or 'N/A' if not found
        """
        import re
        if not manifold:
            return 'N/A'

        match = re.search(r'KJ(\d{1,2})', manifold)
        return f"KJ{match.group(1)}" if match else 'N/A'

    def _build_valve_diagram_name(self, kj_name: str, valve_position: str) -> str:
        """
        Build electrical diagram valve nomenclature.

        Args:
            kj_name: Controls manifold name (e.g., 'KJ1')
            valve_position: Valve position (e.g., '1A', '2B')

        Returns:
            Diagram name (e.g., '=KJ1-QMB1') or 'N/A'
        """
        if kj_name == 'N/A' or not valve_position or valve_position == 'N/A':
            return 'N/A'

        # Extract valve number from '1A' -> '1' or '2B' -> '2'
        valve_num = valve_position.rstrip('AB')

        # Prefix with single quote to force Excel to treat as text (not formula)
        return f"'={kj_name}-QMB{valve_num}"

    def _write_transitions_section(self, ws, row_num: int, id_counter: int, transitions_data: Dict[str, Any],
                                    digital_inputs_data: Dict[str, Any], model_data: Dict[str, Any],
                                    actuator_groups_data: Dict[str, Any], valve_mappings_data: Dict[str, Any]) -> tuple:
        """
        Write transitions section with Sequence Header, Fixed State, Wait Conditions, Actions, and End Of Sequence rows.

        Args:
            ws: Worksheet
            row_num: Current row number
            id_counter: Current ID counter
            transitions_data: Transitions data
            digital_inputs_data: Digital inputs data
            model_data: Model data for extracting sequence style and sequences
            actuator_groups_data: Actuator groups data for MM descriptions
            valve_mappings_data: Valve mappings data for valve info

        Returns:
            Tuple of (updated row_num, updated id_counter)
        """
        if not transitions_data or not transitions_data.get('transitions'):
            return row_num, id_counter

        # Extract model name for Style column (e.g., "EmStatesAndSequences_R2S" -> "R2S")
        routine_name = model_data.get('routine_name', '') if model_data else ''
        sequence_style = routine_name.split('_')[-1] if '_' in routine_name else routine_name

        # Build MM description mapping
        mm_to_description = {}
        if actuator_groups_data and actuator_groups_data.get('actuator_groups'):
            for group in actuator_groups_data['actuator_groups']:
                mm_to_description[group['tag_name']] = group['description']

        # Build valve mapping dictionary
        mm_to_valve = {}
        if valve_mappings_data and valve_mappings_data.get('valve_mappings'):
            mm_to_valve = valve_mappings_data['valve_mappings']

        # Get sequences from model_data
        sequences = model_data.get('sequences', []) if model_data else []

        # IMPROVED: Build sequence lookup dictionary for faster access and better matching
        sequences_by_index = {seq['sequence_index']: seq for seq in sequences}

        # Write Sequence Header row at the beginning
        row_data = [
            id_counter, sequence_style, 'Sequence Header',
            None, None, f"{sequence_style} Sequence",
            None, None, None, None, None, None,
            None, None, None, None, None
        ]
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.fill = self.data_fill
            cell.font = self.data_font
        row_num += 1
        id_counter += 1

        # Process each transition
        for transition_idx, transition in enumerate(transitions_data['transitions']):
            transition_index = transition['transition_index']
            transition_name = transition.get('transition_name', f"State{transition_index}")

            # Parse transition_name to extract Row Type and state name
            if 'Transition State' in transition_name:
                row_type = 'Transition State'
                state_name = transition_name.split(' - ', 1)[1] if ' - ' in transition_name else transition_name
            elif 'Fixed State' in transition_name:
                row_type = 'Fixed State'
                state_name = transition_name.split(' - ', 1)[1] if ' - ' in transition_name else transition_name
            else:
                # Default to Fixed State
                row_type = 'Fixed State'
                state_name = transition_name

            # Write Fixed State / Transition State row
            row_data = [
                id_counter,                      # ID
                sequence_style,                  # Style (e.g., "R2S")
                row_type,                        # Row Type (Fixed State or Transition State)
                None,                            # Actor Type (empty)
                None,                            # Actor/Unit (empty - filled manually)
                state_name,                      # Custom Description (state name)
                None,                            # Custom Duration (empty)
                None,                            # Standard Action/Status (empty)
                None,                            # Standard Duration (empty)
                None,                            # Actor/Group Description (empty)
                None,                            # Actor/Group Name (empty)
                None,                            # Valve Name (empty)
                None, None, None, None, None     # Timing columns (empty)
            ]

            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.fill = self.fixed_state_fill
                cell.font = self.fixed_state_font

            row_num += 1
            id_counter += 1

            # Write Wait Conditions rows (expand permissions)
            for permission in transition['permissions']:
                wait_conditions = self._expand_permission_to_wait_conditions(
                    permission, digital_inputs_data
                )

                for wait_condition in wait_conditions:
                    row_data = [
                        id_counter,                                    # ID
                        sequence_style,                                # Style
                        'Wait Conditions',                             # Row Type
                        wait_condition.get('actor_type'),              # Actor Type
                        None,                                          # Actor/Unit (empty - filled manually)
                        wait_condition.get('custom_desc'),             # Custom Description
                        wait_condition.get('custom_duration'),         # Custom Duration
                        wait_condition.get('status'),                  # Standard Action/Status
                        wait_condition.get('standard_duration', 0.0),  # Standard Duration
                        wait_condition.get('group_desc'),              # Actor/Group Description
                        wait_condition.get('group_name'),              # Actor/Group Name
                        None,                                          # Valve Name (empty)
                        None, None, None, None, None                   # Timing columns (empty)
                    ]

                    for col_num, value in enumerate(row_data, 1):
                        cell = ws.cell(row=row_num, column=col_num, value=value)
                        cell.fill = self.data_fill
                        cell.font = self.data_font

                    row_num += 1
                    id_counter += 1

            # IMPROVED: Multi-strategy fallback for sequence-to-transition matching
            matching_sequence = None

            # Strategy 1: Direct index match
            if transition_index in sequences_by_index:
                matching_sequence = sequences_by_index[transition_index]

            # Strategy 2: Try index + 1 (common pattern where transition 0 -> sequence 1)
            elif (transition_index + 1) in sequences_by_index:
                matching_sequence = sequences_by_index[transition_index + 1]
                if self.debug if hasattr(self, 'debug') else False:
                    logger.debug(f"Matched transition {transition_index} to sequence {transition_index + 1} (offset by 1)")

            # Strategy 3: Use sequence at same position in list (fallback for irregular indexing)
            elif transition_idx < len(sequences):
                matching_sequence = sequences[transition_idx]
                if self.debug if hasattr(self, 'debug') else False:
                    logger.debug(f"Matched transition {transition_index} to sequence at position {transition_idx} (by position)")

            # Strategy 4: Log warning if no match found
            if not matching_sequence:
                logger.warning(f"⚠️ Warning: No sequence found for transition {transition_index} ({transition_name})")
            else:
                # ═══════════════════════════════════════════════════════════
                # Write Transition State row (from sequence_name)
                # ═══════════════════════════════════════════════════════════
                transition_state_name = matching_sequence.get('sequence_name', '')

                row_data = [
                    id_counter,                      # ID
                    sequence_style,                  # Style (e.g., "R2S")
                    'Transition State',              # Row Type
                    None,                            # Actor Type (empty)
                    None,                            # Actor/Unit (empty - filled manually)
                    transition_state_name,           # Custom Description (sequence name)
                    None,                            # Custom Duration (empty)
                    None,                            # Standard Action/Status (empty)
                    None,                            # Standard Duration (empty)
                    None,                            # Actor/Group Description (empty)
                    None,                            # Actor/Group Name (empty)
                    None,                            # Valve Name (empty)
                    None, None, None, None, None     # Timing columns (empty)
                ]

                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.fill = self.transition_state_fill
                    cell.font = self.transition_state_font

                row_num += 1
                id_counter += 1

                # Now write action rows for this sequence (Step1, Step2, Step3)
                row_num, id_counter = self._write_sequence_actions(
                    ws, row_num, id_counter, matching_sequence, sequence_style,
                    mm_to_description, mm_to_valve
                )

        # Write End Of Sequence row
        row_data = [
            id_counter, sequence_style, 'Sequence Header',
            None, None, 'End Of Sequence',
            None, None, None, None, None, None,
            None, None, None, None, None
        ]
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.fill = self.data_fill
            cell.font = self.data_font
        row_num += 1
        id_counter += 1

        return row_num, id_counter

    def _write_sequence_actions(self, ws, row_num: int, id_counter: int, sequence: Dict[str, Any],
                                 sequence_style: str, mm_to_description: Dict[str, str],
                                 mm_to_valve: Dict[str, Dict[str, str]]) -> tuple:
        """
        Write sequence action rows with Step1, Step2, Step3 row types.

        Args:
            ws: Worksheet
            row_num: Current row number
            id_counter: Current ID counter
            sequence: Sequence data with steps and actions
            sequence_style: Style name (e.g., "R2S")
            mm_to_description: MM group descriptions
            mm_to_valve: Valve mappings

        Returns:
            Tuple of (updated row_num, updated id_counter)
        """
        # Process each step with step number
        for step_idx, step in enumerate(sequence.get('steps', []), 1):
            step_name = f"Step{step_idx}"  # "Step1", "Step2", "Step3"

            for action in step.get('actions', []):
                mm_number = action.get('mm_number', '')
                state = action.get('state', '')
                state_formatted = self._format_state_robust(state)

                # Get MM group description
                mm_description = mm_to_description.get(mm_number, '')

                # Calculate valve name using helper method
                valve_name = self._calculate_valve_name(state_formatted, mm_number, mm_to_valve)

                # Write one row per actuator
                for actuator in action.get('actuators', []):
                    actuator_name = actuator['description']

                    # Format actuator name with = and -
                    actor_group_name = self._format_actor_group_name(actuator_name)

                    row_data = [
                        id_counter,                      # ID
                        sequence_style,                  # Style (e.g., "R2S")
                        step_name,                       # Row Type (Step1, Step2, etc.)
                        'CylinderUnits',                 # Actor Type
                        None,                            # Actor/Unit (empty - filled manually)
                        None,                            # Custom Description (empty)
                        None,                            # Custom Duration (empty)
                        state_formatted,                 # Standard Action/Status (TO WORK/HOME)
                        2.0,                             # Standard Duration
                        mm_description,                  # Actor/Group Description
                        actor_group_name,                # Actor/Group Name (=MM1-MMB1)
                        valve_name,                      # Valve Name
                        None, None, None, None, None     # Timing columns (empty)
                    ]

                    for col_num, value in enumerate(row_data, 1):
                        cell = ws.cell(row=row_num, column=col_num, value=value)
                        cell.fill = self.data_fill
                        cell.font = self.data_font

                    row_num += 1
                    id_counter += 1

        return row_num, id_counter

    def _expand_permission_to_wait_conditions(self, permission: Dict[str, Any],
                                               digital_inputs_data: Dict[str, Any]) -> List[Dict[str, Any]]:
        """
        IMPROVED: Expand a transition permission into wait condition rows using 5-pattern matching.

        Args:
            permission: Permission dictionary with permission_value and comment
            digital_inputs_data: Digital inputs data

        Returns:
            List of wait condition dictionaries
        """
        import re

        permission_value = permission.get('permission_value', '')
        comment = permission.get('comment', '')

        wait_conditions = []

        # Pattern 1: All Parts Present/Valid - expand to individual sensors
        if 'AllParts' in permission_value:
            if digital_inputs_data and digital_inputs_data.get('digital_inputs'):
                # Sort sensors by part assignment for consistent output
                sensors_list = []
                for digital_input in digital_inputs_data['digital_inputs']:
                    part_assignment = digital_input.get('part_assignment', 'N/A')

                    # Only include sensors with part assignments and starting with BG
                    if part_assignment == 'N/A' or not digital_input['tag_name'].startswith('BG'):
                        continue

                    sensors_list.append((part_assignment, digital_input))

                # Sort by part assignment using robust part name sorting
                sensors_list.sort(key=lambda x: self._sort_part_name(x[0]))

                for part_assignment, digital_input in sensors_list:
                    tag_name = digital_input['tag_name']
                    description = digital_input.get('description', '')

                    # Format sensor name with = and -
                    sensor_name_formatted = self._format_actor_group_name(tag_name)

                    wait_conditions.append({
                        'actor_type': 'SensorUnits',
                        'actor_unit': None,
                        'custom_desc': None,
                        'custom_duration': None,
                        'status': 'ON',
                        'standard_duration': 0.0,
                        'group_desc': description or part_assignment,
                        'group_name': sensor_name_formatted
                    })

        # Pattern 2: Specific Part Present (NEW) - e.g., Part1Present, Part2Valid
        elif re.search(r'Part\d+(?:Present|Valid|Detected)', permission_value, re.IGNORECASE):
            # Extract part number
            match = re.search(r'Part(\d+)', permission_value, re.IGNORECASE)
            if match and digital_inputs_data and digital_inputs_data.get('digital_inputs'):
                part_num = match.group(1)
                target_part = f"Part{part_num}"

                # Find sensors for this specific part
                for digital_input in digital_inputs_data['digital_inputs']:
                    part_assignment = digital_input.get('part_assignment', 'N/A')

                    if part_assignment == target_part and digital_input['tag_name'].startswith('BG'):
                        tag_name = digital_input['tag_name']
                        description = digital_input.get('description', '')

                        # Format sensor name with = and -
                        sensor_name_formatted = self._format_actor_group_name(tag_name)

                        wait_conditions.append({
                            'actor_type': 'SensorUnits',
                            'actor_unit': None,
                            'custom_desc': None,
                            'custom_duration': None,
                            'status': 'ON',
                            'standard_duration': 0.0,
                            'group_desc': description or part_assignment,
                            'group_name': sensor_name_formatted
                        })

        # Pattern 3: Timer/Delay conditions (NEW) - extract duration
        elif re.search(r'(timer|delay|wait)', permission_value, re.IGNORECASE):
            duration = self._extract_duration(permission_value) or self._extract_duration(comment)

            wait_conditions.append({
                'actor_type': 'Timers',
                'actor_unit': None,
                'custom_desc': comment or permission_value,
                'custom_duration': duration,
                'status': None,
                'group_desc': None,
                'group_name': None
            })

        # Pattern 4: Operators - manual operator actions
        elif 'Operator' in comment or 'operator' in comment.lower() or 'Load' in comment or 'Leave' in comment:
            # Extract operator number
            operator_match = re.search(r'Operator\s*(\d+)', comment, re.IGNORECASE)
            operator_unit = f"Operator {operator_match.group(1)}" if operator_match else "Operator 1"

            # Extract duration if present
            duration = self._extract_duration(comment) or self._extract_duration(permission_value)

            wait_conditions.append({
                'actor_type': 'Operators',
                'actor_unit': None,
                'custom_desc': comment,
                'custom_duration': duration,
                'status': None,
                'standard_duration': None,
                'group_desc': None,
                'group_name': None
            })

        # Pattern 5: Robot status - show as robot wait condition
        elif '.Rbt.' in permission_value or 'Robot' in comment or 'robot' in comment.lower():
            # Extract robot identifier using helper method
            robot_unit = self._extract_robot_unit(permission_value, comment)
            custom_desc = comment or permission_value

            # Extract duration if present
            duration = self._extract_duration(comment) or self._extract_duration(permission_value)

            wait_conditions.append({
                'actor_type': 'Robots',
                'actor_unit': robot_unit,
                'custom_desc': custom_desc,
                'custom_duration': duration,
                'status': None,
                'standard_duration': None,
                'group_desc': None,
                'group_name': None
            })

        # Pattern 5: Cylinder position checks (NEW) - e.g., MM1_Home, MM2_Work
        elif re.search(r'MM\d+[._]?(Home|Work|stsAt)', permission_value, re.IGNORECASE):
            # Extract MM number and state
            match = re.search(r'MM(\d+)[._]?(Home|Work|stsAt(Home|Work))', permission_value, re.IGNORECASE)
            if match:
                mm_num = match.group(1)
                state_raw = match.group(2)

                # Determine state
                if 'Work' in state_raw:
                    state = 'Work'
                else:
                    state = 'Home'

                wait_conditions.append({
                    'actor_type': 'CylinderUnits',
                    'actor_unit': None,
                    'custom_desc': None,
                    'custom_duration': None,
                    'status': self._format_state_robust(state),
                    'group_desc': f"MM{mm_num} Group",
                    'group_name': f"MM{mm_num}"
                })
            else:
                # Fallback to generic if regex fails
                wait_conditions.append({
                    'actor_type': None,
                    'actor_unit': None,
                    'custom_desc': permission_value,
                    'custom_duration': None,
                    'status': None,
                    'group_desc': comment,
                    'group_name': None
                })

        # Pattern 6: Other conditions - show as generic wait condition
        else:
            wait_conditions.append({
                'actor_type': None,
                'actor_unit': None,
                'custom_desc': permission_value,
                'custom_duration': None,
                'status': None,
                'group_desc': comment,
                'group_name': None
            })

        return wait_conditions

    def _format_state_robust(self, state_value):
        """
        Robustly format cylinder state values for Actions (uses "TO" prefix).

        Args:
            state_value: Raw state value from data (may be None, empty, or uppercase/lowercase)

        Returns:
            Formatted state string like "TO HOME" or "TO WORK"
        """
        if not state_value:
            return "TO HOME"

        # Normalize: strip whitespace and convert to uppercase
        state_normalized = str(state_value).strip().upper()

        # If already has "TO", keep as is
        if state_normalized.startswith("TO "):
            return state_normalized

        # If empty after normalization, default to HOME
        if not state_normalized:
            return "TO HOME"

        # Otherwise, add "TO" prefix
        return f"TO {state_normalized}"

    def _format_start_condition_state(self, state_value):
        """
        Format cylinder state values for Start Conditions (uses "AT" prefix).

        Args:
            state_value: Raw state value from data (may be None, empty, or uppercase/lowercase)

        Returns:
            Formatted state string like "AT HOME" or "AT WORK"
        """
        if not state_value:
            return "AT HOME"

        # Normalize: strip whitespace and convert to uppercase
        state_normalized = str(state_value).strip().upper()

        # Remove "TO" prefix if present (from overrides)
        if state_normalized.startswith("TO "):
            state_normalized = state_normalized[3:].strip()

        # If already has "AT", keep as is
        if state_normalized.startswith("AT "):
            return state_normalized

        # If empty after normalization, default to HOME
        if not state_normalized:
            return "AT HOME"

        # Otherwise, add "AT" prefix
        return f"AT {state_normalized}"

    def _select_valve_position(self, state: str, valve_info: Dict[str, str]) -> str:
        """
        Select the correct valve position based on cylinder state.

        Args:
            state: Cylinder state (e.g., "TO HOME", "TO WORK")
            valve_info: Dictionary with 'valve_work' and 'valve_home' keys

        Returns:
            Valve position string (e.g., "1A", "2B") or empty string
        """
        if not valve_info:
            return ''

        # Normalize state for comparison
        state_normalized = str(state).upper()

        # Check for WORK state
        if 'WORK' in state_normalized:
            return valve_info.get('valve_work', '')

        # Default to HOME (covers "HOME", "TO HOME", or any other state)
        return valve_info.get('valve_home', '')

    def _calculate_valve_name(self, state: str, mm_number: str,
                              mm_to_valve: Dict[str, Dict[str, str]]) -> str:
        """
        Calculate the complete valve name for a cylinder in a given state.

        Args:
            state: Cylinder state
            mm_number: MM group tag (e.g., "MM1")
            mm_to_valve: Valve mappings dictionary

        Returns:
            Valve diagram name (e.g., "'=KJ1-QMB1") or None
        """
        if mm_number not in mm_to_valve:
            return None

        valve_info = mm_to_valve[mm_number]
        manifold = valve_info.get('manifold', '')

        if not manifold:
            return None

        # Select correct valve position
        valve_position = self._select_valve_position(state, valve_info)

        if not valve_position or valve_position == 'N/A':
            return None

        # Build diagram name
        kj_name = self._extract_kj_name(manifold)
        return self._build_valve_diagram_name(kj_name, valve_position)

    def _extract_duration(self, text: str) -> str:
        """
        Extract duration/timing value from text.

        Args:
            text: Text containing duration (e.g., "Timer > 500ms", "Delay 2s")

        Returns:
            Formatted duration string or empty string
        """
        import re

        # Look for patterns like "500ms", "2s", "1.5s"
        match = re.search(r'(\d+(?:\.\d+)?)\s*(ms|s|sec|seconds?)', text, re.IGNORECASE)
        if match:
            value = match.group(1)
            unit = match.group(2).lower()

            # Normalize unit
            if unit in ['s', 'sec', 'second', 'seconds']:
                return f"{value}s"
            else:  # ms
                return f"{value}ms"

        return ''

    def _extract_robot_unit(self, permission_value: str, comment: str) -> str:
        """
        Extract robot unit identifier from permission or comment.

        Args:
            permission_value: Permission value string
            comment: Comment string

        Returns:
            Robot unit identifier (e.g., "Robot 1", "Robot 2")
        """
        import re

        # Try to find robot number in permission value
        match = re.search(r'Rbt[._]?(\d+)', permission_value, re.IGNORECASE)
        if match:
            return f"Robot {match.group(1)}"

        # Try to find in comment
        match = re.search(r'Robot\s*(\d+)', comment, re.IGNORECASE)
        if match:
            return f"Robot {match.group(1)}"

        return 'Robot 1'  # Default

    def _sort_part_name(self, part_name: str) -> tuple:
        """
        Generate sort key for part names to handle various formats.

        Args:
            part_name: Part name (e.g., "Part1", "Part12", "PartLeft")

        Returns:
            Sort key tuple (priority, numeric_value, string_value)
        """
        import re

        # Try to extract numeric part
        match = re.match(r'Part(\d+)([A-Za-z]*)', part_name)
        if match:
            num = int(match.group(1))
            suffix = match.group(2)
            return (0, num, suffix)  # Numbered parts first, sorted numerically

        # Non-numeric parts (e.g., "PartLeft", "PartRight")
        return (1, 0, part_name)  # Then alphabetically

    def _format_actor_group_name(self, name: str) -> str:
        """
        Format actor/group name: "MM1_MMB1" → "'=MM1-MMB1"

        Args:
            name: Original name with underscore separator

        Returns:
            Formatted name with '= prefix and - separator, or None if empty
        """
        if not name:
            return None

        # Replace underscore with hyphen
        formatted = name.replace('_', '-')

        # Add '= prefix (single quote + equals) to force Excel to treat as text (not formula)
        return f"'={formatted}"

    def _adjust_column_widths(self, ws):
        """
        Auto-adjust column widths based on content.

        Args:
            ws: Worksheet object
        """
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max_length + DEFAULT_COLUMN_PADDING, MAX_COLUMN_WIDTH)
            ws.column_dimensions[column_letter].width = adjusted_width
